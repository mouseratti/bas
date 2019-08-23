# https://openpyxl.readthedocs.io/en/stable/tutorial.html
import openpyxl
from datetime import datetime
from random import  randint
FILENAME = 'workbook.xlsx'
# Create new xlsx
wb = openpyxl.Workbook()
# create new sheet
wb.create_sheet("new worksheeeet", 1)

# move to sheet by its index
wb.active = 1

# create link to worksheet
ws = wb.active

ws['A1'] = 'Table 1'
ws.append(('position', 'speed, km\h', 'time, h', 'distance, km'))
for row in range(1, 11):
    speed, time = randint(20, 200), randint(1,4)
    distance = speed * time
    ws.append((row, speed, time, distance))

ws['A3'] = datetime.now()
wb.save(FILENAME)

# load from file
workbook = openpyxl.load_workbook(FILENAME)
print(workbook.sheetnames)
